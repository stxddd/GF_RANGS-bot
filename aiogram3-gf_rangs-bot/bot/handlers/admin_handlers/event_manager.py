import re
from datetime import datetime
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from utils.decorators.admin_required import admin_required
from bot.db.events.dao import EventDAO, RoleDAO, UserEventRoleDAO
from bot.db.users.dao import UserDAO
from bot.templates.message_templates import event_message
from bot.templates.kb_templates import add_event_text, view_all_events_text,view_all_users_text
from bot.templates.message_templates import (
    ENTER_EVENT_NAME_MESSAGE,
    EMPTY_EVENT_NAME_MESSAGE,
    EVENT_ADDED_SUCCESS_MESSAGE,
    EVENT_ADD_ERROR_MESSAGE,
    NO_EVENTS_MESSAGE,
    ALL_EVENTS_MESSAGE,
    NO_EVENT_FOUND_MESSAGE,
    CONFIRM_DELETE_EVENT_MESSAGE,
    EVENT_DELETED_SUCCESS_MESSAGE,
    EVENT_DELETE_ERROR_MESSAGE,
)
from bot.kb.events_kb import get_events_for_edit, yes_or_not_delete_event_keyboard
from bot.kb.users_kb import get_users_to_edit
from utils.decorators.admin_required import admin_required

router = Router()


class AddEventState(StatesGroup):
    waiting_for_name = State()

class RenameEventState(StatesGroup):
    waiting_for_new_name = State()


EDIT_PAGE_PATTERN = r"^edit_page_(\d+)$"
GET_CLIENT_TO_EDIT_PATTERN = r"^get_event_to_edit_(\d+)$"

PREPARE_TO_DELETE_EVENT_PATTERN = r"^prepare_to_delete_event_(\d+)$"
DELETE_EVENT_PATTERN = r"^delete_event_(\d+)$"


@router.message(F.text == add_event_text)
@admin_required
async def start_add_event(message: Message, state: FSMContext):
    await state.set_state(AddEventState.waiting_for_name)
    await message.answer(ENTER_EVENT_NAME_MESSAGE)


@router.message(AddEventState.waiting_for_name)
@admin_required
async def add_event(message: Message, state: FSMContext):
    name = message.text.strip()

    if not name:
        return await message.answer(EMPTY_EVENT_NAME_MESSAGE)

    result = await EventDAO.add(name=name, visibility = True)
    if result:
        await message.answer(EVENT_ADDED_SUCCESS_MESSAGE.format(name=name))
    else:
        await message.answer(EVENT_ADD_ERROR_MESSAGE)

    await state.clear()


@router.callback_query(F.data.regexp(EDIT_PAGE_PATTERN))
@admin_required
async def handle_pagination(callback: CallbackQuery):
    await callback.answer()

    match = re.match(EDIT_PAGE_PATTERN, callback.data)
    page = int(match.group(1))

    events = await EventDAO.find_all()

    if not events:
        return await callback.message.answer(NO_EVENTS_MESSAGE)

    await callback.message.edit_reply_markup(
        reply_markup=await get_events_for_edit(tg_id = callback.from_user.id, events=events, page=page)
    )


@router.message(F.text == view_all_events_text)
@admin_required
async def handle_get_events(message: Message):
    events = await EventDAO.find_all()
    if not events:
        return await message.answer(NO_EVENTS_MESSAGE)

    await message.answer(
        ALL_EVENTS_MESSAGE,
        reply_markup=await get_events_for_edit(tg_id = message.from_user.id,events=events, page=1),
    )


@router.callback_query(F.data.regexp(GET_CLIENT_TO_EDIT_PATTERN))
@admin_required
async def handle_get_event(callback: CallbackQuery):
    await callback.answer()

    match = re.match(GET_CLIENT_TO_EDIT_PATTERN, callback.data)
    event_id = int(match.group(1))

    current_event = await EventDAO.find_by_id(model_id=event_id)
    if not current_event:
        return await callback.message.answer(NO_EVENT_FOUND_MESSAGE)

    roles = await RoleDAO.find_all(event_id=current_event.id)
    await callback.message.answer(
        event_message(roles, current_event),
        reply_markup=await get_users_to_edit(event_id=event_id),
    )


@router.callback_query(F.data.startswith("toggle_visibility:"))
@admin_required
async def toggle_event_visibility(callback: CallbackQuery):
    await callback.answer()
    event_id = int(callback.data.split(":")[1])
    event = await EventDAO.find_by_id(event_id)
    event = await EventDAO.update(model_id = event_id, visibility = not(event.visibility))


@router.callback_query(F.data.regexp(PREPARE_TO_DELETE_EVENT_PATTERN))
@admin_required
async def handle_prepare_to_delete_event(callback: CallbackQuery, state: FSMContext):
    await callback.answer()

    match = re.match(PREPARE_TO_DELETE_EVENT_PATTERN, callback.data)
    event_id = int(match.group(1))

    event = await EventDAO.find_by_id(event_id)
    if not event:
        return await callback.message.answer(NO_EVENT_FOUND_MESSAGE)

    return await callback.message.answer(
        CONFIRM_DELETE_EVENT_MESSAGE.format(name=event.name),
        reply_markup=yes_or_not_delete_event_keyboard(event_id=event_id),
    )


@router.callback_query(F.data.regexp(DELETE_EVENT_PATTERN))
@admin_required
async def handle_delete_event(callback: CallbackQuery, state: FSMContext):
    await callback.answer()

    match = re.match(DELETE_EVENT_PATTERN, callback.data)
    event_id = int(match.group(1))

    event = await EventDAO.find_by_id(event_id)
    if not event:
        return await callback.message.answer(NO_EVENT_FOUND_MESSAGE)

    deleted = await EventDAO.delete(id=event_id)
    if not deleted:
        return await callback.message.answer(EVENT_DELETE_ERROR_MESSAGE)

    return await callback.message.answer(EVENT_DELETED_SUCCESS_MESSAGE)


@router.callback_query(F.data.startswith("rename_event:"))
@admin_required
async def handle_rename_event(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    event_id = int(callback.data.split(":")[1])
    event = await EventDAO.find_by_id(event_id)

    if not event:
        return await callback.message.answer(NO_EVENT_FOUND_MESSAGE)

    await state.update_data(event_id=event_id)
    await state.set_state(RenameEventState.waiting_for_new_name)
    await callback.message.answer(f"Введите новое название для мероприятия: {event.name}")


@router.message(RenameEventState.waiting_for_new_name)
@admin_required
async def process_rename_event(message: Message, state: FSMContext):
    new_name = message.text.strip()

    if not new_name:
        return await message.answer("Название не может быть пустым!")

    data = await state.get_data()
    event_id = data.get("event_id")

    updated = await EventDAO.update(model_id=event_id, name=new_name)
    if not updated:
        return await message.answer("Ошибка при переименовании мероприятия.")

    await message.answer(f"Мероприятие успешно переименовано в: {new_name}")
    await state.clear()


@router.message(F.text == view_all_users_text)
@admin_required
async def get_data(message: Message):
    users = await UserDAO.find_all()

    user_points = []
    for user in users:
        total_points = await UserDAO.get_total_points_by_user_id(user.id)
        user_points.append((user.fullname, total_points))

    user_points.sort(key=lambda x: x[1], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Баллы за мероприятия"

    ws["A1"] = "ФИО"
    ws["B1"] = "Баллов"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in ("A", "B"):
        ws[f"{col}1"].font = header_font
        ws[f"{col}1"].fill = header_fill
        ws[f"{col}1"].alignment = alignment
        ws[f"{col}1"].border = thin_border

    row = 2
    for fullname, points in user_points:
        ws[f"A{row}"] = fullname
        ws[f"B{row}"] = points

        for col in ("A", "B"):
            ws[f"{col}{row}"].alignment = alignment
            ws[f"{col}{row}"].border = thin_border
        row += 1

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M")
    filename = f"Итог_на_{timestamp}.xlsx"

    await message.answer_document(
        document=BufferedInputFile(output.getvalue(), filename)
    )